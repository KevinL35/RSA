from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any
from uuid import UUID

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as excel_serial_to_datetime
from supabase import Client

from app.modules.tasks import state_machine
from app.modules.tasks.listing import enrich_task_for_task_center

MAX_IMPORT_ROWS = 20000
MAX_FILE_BYTES = 10 * 1024 * 1024


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_header_cell(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return ""
    return str(val).strip()


def _find_time_col_index(headers: list[str]) -> int | None:
    exact = (
        "reviewed_at",
        "发布时间",
        "评论时间",
        "时间",
        "日期",
        "date",
        "time",
        "review time",
        "reviewed at",
    )
    lowered = [h.strip().lower() for h in headers]
    for e in exact:
        el = e.lower()
        for i, sl in enumerate(lowered):
            if sl == el:
                return i
    for i, h in enumerate(headers):
        s = h.strip()
        if not s:
            continue
        if "评论时间" in s or "发布时间" in s:
            return i
    for i, h in enumerate(headers):
        s = h.strip()
        if s in ("时间", "日期"):
            return i
    for i, sl in enumerate(lowered):
        if sl.endswith(" time") or sl.endswith("_at") or sl == "at":
            return i
    return None


def _find_text_col_index(headers: list[str]) -> int | None:
    exact = (
        "评论",
        "正文",
        "内容",
        "raw_text",
        "text",
        "content",
        "review",
        "comment",
        "body",
        "message",
    )
    lowered = [h.strip().lower() for h in headers]
    for e in exact:
        el = e.lower()
        for i, sl in enumerate(lowered):
            if sl == el:
                return i
    for i, h in enumerate(headers):
        s = h.strip()
        if "评论" in s and "时间" not in s:
            return i
    return None


def _cell_to_reviewed_at_iso(val: object) -> str | None:
    if val is None:
        return None
    if isinstance(val, str) and not val.strip():
        return None
    if isinstance(val, datetime):
        dt = val
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt.isoformat()
    if isinstance(val, date):
        dt = datetime.combine(val, datetime.min.time(), tzinfo=timezone.utc)
        return dt.isoformat()
    if isinstance(val, (int, float)):
        try:
            d = excel_serial_to_datetime(val)
            return _cell_to_reviewed_at_iso(d)
        except Exception:  # noqa: BLE001
            return None
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(s[:19], fmt).replace(tzinfo=timezone.utc)
                return dt.isoformat()
            except ValueError:
                continue
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc).isoformat()
        except ValueError:
            return None
    return None


def _cell_to_raw_text(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_review_grid(headers: list[str], data_rows: list[tuple[object, ...]]) -> list[dict[str, Any]]:
    ti = _find_time_col_index(headers)
    ci = _find_text_col_index(headers)
    if ti is None or ci is None:
        raise ValueError(
            "未识别表头：请包含「时间」与「评论」列（或英文 date / reviewed_at、text / raw_text 等）",
        )
    if ti == ci:
        raise ValueError("时间与评论不能为同一列")

    out: list[dict[str, Any]] = []
    for row in data_rows:
        ti_v = row[ti] if ti < len(row) else None
        ci_v = row[ci] if ci < len(row) else None
        raw_text = _cell_to_raw_text(ci_v)
        if not raw_text:
            continue
        reviewed_at = _cell_to_reviewed_at_iso(ti_v)
        out.append({"raw_text": raw_text, "reviewed_at": reviewed_at})
        if len(out) > MAX_IMPORT_ROWS:
            raise ValueError(f"评论行数超过上限 {MAX_IMPORT_ROWS}")

    if not out:
        raise ValueError("没有有效数据行：请至少提供一条非空评论")
    return out


def _parse_xlsx_review_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"无法读取 .xlsx：{e!s}") from e
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel 无活动工作表")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as e:
            raise ValueError("Excel 无内容") from e
        headers = [_normalize_header_cell(c) for c in header_row]
        data_rows: list[tuple[object, ...]] = []
        for row in rows_iter:
            if row is None:
                continue
            data_rows.append(tuple(row))
        return _parse_review_grid(headers, data_rows)
    finally:
        wb.close()


def _parse_xls_review_rows(content: bytes) -> list[dict[str, Any]]:
    import xlrd  # type: ignore[import-untyped]

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"无法读取 .xls：{e!s}") from e
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Excel 无内容")

    def cell_value(r: int, c: int) -> object:
        cell = sheet.cell(r, c)
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, book.datemode)
            except Exception:  # noqa: BLE001
                return cell.value
        return cell.value

    headers = [_normalize_header_cell(cell_value(0, c)) for c in range(sheet.ncols)]
    data_rows: list[tuple[object, ...]] = []
    for r in range(1, sheet.nrows):
        data_rows.append(tuple(cell_value(r, c) for c in range(sheet.ncols)))
    return _parse_review_grid(headers, data_rows)


def parse_excel_review_rows(content: bytes, filename: str = "") -> list[dict[str, Any]]:
    """从 Excel 首行表头解析「时间」「评论」列；支持 .xlsx（openpyxl）与 .xls（xlrd）。"""
    if not content:
        raise ValueError("文件为空")
    fn = (filename or "").lower()
    is_xls = fn.endswith(".xls") and not fn.endswith(".xlsx")
    if is_xls:
        return _parse_xls_review_rows(content)
    return _parse_xlsx_review_rows(content)


def run_import_reviews_from_excel(
    sb: Client,
    task_id: UUID,
    content: bytes,
    filename: str = "",
) -> dict:
    """将任务置为 running（若原为 pending）、替换 reviews 为 Excel 内容；与 fetch-reviews 状态规则一致。"""
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"文件过大（上限 {MAX_FILE_BYTES // (1024 * 1024)}MB）",
        )

    try:
        parsed = parse_excel_review_rows(content, filename=filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    res = (
        sb.table("insight_tasks")
        .select("*")
        .eq("id", str(task_id))
        .limit(1)
        .execute()
    )
    rows = res.data or []
    if not rows:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = rows[0]
    status = task["status"]
    if status in ("success", "cancelled"):
        raise HTTPException(
            status_code=409,
            detail=f"任务状态为 {status}，不可导入评论",
        )
    if status == "failed":
        raise HTTPException(
            status_code=400,
            detail="任务已失败：请先 POST /api/v1/insight-tasks/{id}/retry 重置为 pending，再导入评论",
        )

    platform = task["platform"]
    product_id = task["product_id"]

    if status == "pending":
        try:
            state_machine.assert_valid_transition("pending", "running")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        up = (
            sb.table("insight_tasks")
            .update(
                {
                    "status": "running",
                    "updated_at": _utc_now_iso(),
                    "error_code": None,
                    "error_message": None,
                    "failure_stage": None,
                },
            )
            .eq("id", str(task_id))
            .execute()
        )
        if not up.data:
            raise HTTPException(status_code=500, detail="无法将任务置为 running")
        task = up.data[0]

    sb.table("reviews").delete().eq("insight_task_id", str(task_id)).execute()

    insert_rows: list[dict[str, Any]] = []
    for r in parsed:
        insert_rows.append(
            {
                "insight_task_id": str(task_id),
                "platform": platform,
                "product_id": product_id,
                "external_review_id": None,
                "raw_text": r["raw_text"],
                "title": None,
                "rating": None,
                "sku": None,
                "reviewed_at": r.get("reviewed_at"),
                "lang": None,
                "extra": None,
            },
        )

    batch = 500
    for i in range(0, len(insert_rows), batch):
        chunk = insert_rows[i : i + batch]
        sb.table("reviews").insert(chunk).execute()

    fresh = (
        sb.table("insight_tasks")
        .select("*")
        .eq("id", str(task_id))
        .limit(1)
        .execute()
    )
    task_row = (fresh.data or [task])[0]
    return {
        "task": enrich_task_for_task_center(task_row),
        "reviews_inserted": len(insert_rows),
    }
