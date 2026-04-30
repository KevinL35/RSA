"""从 Excel 解析词典行并合并进 overlay（表头支持中英文）。"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .taxonomy_yaml import SIX_WAY_DIMENSION_ORDER

MAX_DICTIONARY_IMPORT_ROWS = 5000
MAX_DICTIONARY_FILE_BYTES = 10 * 1024 * 1024


def _norm_header(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip().lower()


def _find_col(headers: list[str], *candidates: str) -> int | None:
    lowered = [_norm_header(h) for h in headers]
    for c in candidates:
        cl = c.lower()
        for i, sl in enumerate(lowered):
            if sl == cl or sl.replace(" ", "_") == cl:
                return i
    return None


def _find_dim_col(headers: list[str]) -> int | None:
    for i, h in enumerate(headers):
        s = _norm_header(h)
        if s in ("dimension_6way", "dim", "dimension"):
            return i
        if s in ("六维", "六维维度", "维度"):
            return i
    return None


def _find_canonical_col(headers: list[str]) -> int | None:
    for i, h in enumerate(headers):
        s = _norm_header(h)
        if s in ("canonical", "keyword", "key"):
            return i
        if s in ("规范词", "关键词"):
            return i
    return None


def _find_aliases_col(headers: list[str]) -> int | None:
    for i, h in enumerate(headers):
        s = _norm_header(h)
        if s in ("aliases", "alias", "synonyms", "synonym"):
            return i
        if s in ("同义词", "别名"):
            return i
    return None


_SPLIT_ALIASES = re.compile(r"[\n;；|,\，]+")


def _split_aliases_cell(val: object) -> list[str]:
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    parts = _SPLIT_ALIASES.split(s)
    return [p.strip() for p in parts if p.strip()]


def parse_dictionary_excel_rows(content: bytes, filename: str = "") -> tuple[list[dict[str, Any]], list[str]]:
    """
    解析首行表头，返回 (行 dict 列表, 错误信息列表)。
    每行 dict: dimension_6way, canonical, aliases, _row（1-based excel row）。
    权重与优先级不在 Excel 中配置；写入 overlay 时使用默认 weight=1.0、priority=50。
    """
    errors: list[str] = []
    if len(content) > MAX_DICTIONARY_FILE_BYTES:
        errors.append(f"文件超过 {MAX_DICTIONARY_FILE_BYTES // (1024 * 1024)}MB")
        return [], errors

    fn = (filename or "").lower()
    if not fn.endswith((".xlsx", ".xls")):
        errors.append("仅支持 .xlsx 或 .xls")
        return [], errors

    if fn.endswith(".xls") and not fn.endswith(".xlsx"):
        errors.append("词典导入请使用 .xlsx（推荐导出后再编辑）")
        return [], errors

    try:
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        errors.append(f"无法读取 Excel：{e!s}")
        return [], errors

    try:
        ws = wb.active
        if ws is None:
            errors.append("Excel 无活动工作表")
            return [], errors
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            errors.append("空表格")
            return [], errors
        headers = [str(c).strip() if c is not None else "" for c in header_row]

        ic_dim = _find_dim_col(headers)
        ic_can = _find_canonical_col(headers)
        ic_als = _find_aliases_col(headers)

        if ic_dim is None:
            errors.append("未找到「六维维度」或 dimension_6way 列")
            return [], errors
        if ic_can is None:
            errors.append("未找到「规范词」或 canonical 列")
            return [], errors
        if ic_als is None:
            errors.append("未找到「同义词」或 aliases 列")
            return [], errors

        out: list[dict[str, Any]] = []
        row_num = 1
        for tup in rows_iter:
            row_num += 1
            if row_num > MAX_DICTIONARY_IMPORT_ROWS + 10:
                errors.append(f"超过最大行数 {MAX_DICTIONARY_IMPORT_ROWS}")
                break
            if not tup or all(v is None or str(v).strip() == "" for v in tup):
                continue
            dim_raw = tup[ic_dim] if ic_dim < len(tup) else None
            can_raw = tup[ic_can] if ic_can < len(tup) else None
            als_raw = tup[ic_als] if ic_als < len(tup) else None
            dim = str(dim_raw or "").strip().lower()
            can = str(can_raw or "").strip()
            if not dim and not can:
                continue
            aliases = _split_aliases_cell(als_raw)

            if dim not in SIX_WAY_DIMENSION_ORDER:
                errors.append(f"第 {row_num} 行：无效维度 {dim_raw!r}（须为 pros/cons/… 之一）")
                continue
            if len(can) < 2:
                errors.append(f"第 {row_num} 行：规范词过短")
                continue

            out.append(
                {
                    "dimension_6way": dim,
                    "canonical": can,
                    "aliases": aliases,
                    "_row": row_num,
                },
            )
    finally:
        wb.close()

    return out, errors
